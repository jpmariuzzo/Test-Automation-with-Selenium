import openpyxl

book = openpyxl.load_workbook("C:\\Users\\jpmar\\PyCharmProjects\\pythonTesting\\pythonSelenium\\PythonDemo.xlsx")
sheet = book.active #Getting control of Active sheet
cell = sheet.cell(row=2, column=2) #Controling the cell
sheet.cell(row=7, column=1).value = "Isabelly Clarice Laura da Paz"
sheet.cell(row=7, column=2).value = "Isabelly"
sheet.cell(row=7, column=3).value = "da Paz"
sheet.cell(row=7, column=4).value = "isabelly-dapaz83@mpv.org.br"

print("Max number of rows: ", sheet.max_row)
print("Max number of columns: ", sheet.max_column)

print(sheet['A7'].value)
print(sheet.cell(row=7, column=4).value)

Dict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Isabelly Clarice Laura da Paz": #Filtering a line
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value


print(Dict)

