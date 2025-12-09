import time

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

service_obj = Service(r"C:\Users\jpmar\PyCharmProjects\pythonTesting\pythonSelenium\chromedriver-win64\chromedriver.exe")

##Configuring Browser
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized") ##Initializing the navigator maximized
driver = webdriver.Chrome(service=service_obj, options=chrome_options)
driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
file_path = r"C:\Users\jpmar\Downloads\download.xlsx"
fruit_name = "Apple"
new_value = "999"

#driver.find_element(By.ID, "downloadButton").click()


def update_excel(filePath, searchTerm, columnName, newValue):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict ={}
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == columnName: #Filtering a line
            Dict["PriceColumn"] = i

    for j in range(1, sheet.max_row+1):
        for k in range(1, sheet.max_column+1):
            if sheet.cell(row=j, column=k).value == searchTerm:
                Dict["FruitRow"] = j
    sheet.cell(row=Dict["FruitRow"], column=Dict["PriceColumn"]).value = newValue
    book.save(filePath)

#Get and edit Excel
update_excel(file_path, fruit_name, "price", new_value)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']") #Type has to be FILE if you want to send a file
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 10)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)") # . to select a class
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text) #We use * to unpack the tuple


priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id") #Getting the Price Column number, generic way
actual_price = driver.find_element(By.XPATH, f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{priceColumn}-undefined']").text

print(actual_price) #Price from Web table
assert actual_price == new_value

time.sleep(3)